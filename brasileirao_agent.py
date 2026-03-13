"""
=============================================================
  AGENTE BRASILEIRAO — KICK MONITOR v2
  Dashboard en vivo + Resumen persistente por partido
=============================================================
INSTALACIÓN:
    pip install requests flask flask-cors

EJECUCIÓN:
    python3 brasileirao_agent.py
    → Abrí http://localhost:5001 en tu browser
=============================================================
"""

import requests, threading, csv, os, time, json
try:
    import cloudscraper as _cs; _session = _cs.create_scraper()
except Exception:
    _session = requests.Session()
from datetime import datetime
import pytz
TZ_AR = pytz.timezone("America/Argentina/Buenos_Aires")
from flask import Flask, jsonify, render_template_string
from flask_cors import CORS

# ─── GOOGLE SHEETS ───────────────────────────────────────────────────────────
try:
    import gspread
    from google.oauth2.service_account import Credentials as _SACredentials
    _GSHEETS_OK = True
except ImportError:
    _GSHEETS_OK = False

SHEETS_SPREADSHEET_ID = "1tkQdTruSgOa2J7aRCJ8T0eTQMBDbtOVaCx9Uup3SmYI"
SHEETS_CREDS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gcp_credentials.json")
SHEETS_CREDS_JSON = os.environ.get("GCP_CREDENTIALS_JSON", "")

def _get_gsheets_client():
    if not _GSHEETS_OK:
        return None
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        if SHEETS_CREDS_JSON:
            import tempfile
            with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
                f.write(SHEETS_CREDS_JSON)
                tmp = f.name
            creds = _SACredentials.from_service_account_file(tmp, scopes=scopes)
            os.unlink(tmp)
        elif os.path.exists(SHEETS_CREDS_FILE):
            creds = _SACredentials.from_service_account_file(SHEETS_CREDS_FILE, scopes=scopes)
        else:
            return None
        return gspread.authorize(creds)
    except Exception as e:
        print(f"[Sheets] Error auth: {e}")
        return None

def update_sheets(resumen: dict):
    """Agrega una fila al Google Sheet con los KPIs del partido."""
    gc = _get_gsheets_client()
    if not gc:
        return
    try:
        sh = gc.open_by_key(SHEETS_SPREADSHEET_ID)
        ws = sh.sheet1
        # Si la hoja está vacía, agregar headers
        if ws.row_count == 0 or ws.cell(1, 1).value != "Fecha":
            headers = ["Fecha", "Canal", "Streamer", "País", "Partido",
                       "Hora Inicio", "Hora Fin", "Duración (min)",
                       "Peak Viewers", "Avg Viewers", "Snapshots", "Min. Vistos Est."]
            ws.append_row(headers)
        row = [
            resumen.get("fecha", ""),
            resumen.get("canal", ""),
            resumen.get("nombre", ""),
            resumen.get("pais", ""),
            resumen.get("titulo", ""),
            resumen.get("hora_inicio", ""),
            resumen.get("hora_fin", ""),
            resumen.get("duracion_min", 0),
            resumen.get("peak_viewers", 0),
            resumen.get("avg_viewers", 0),
            resumen.get("total_snapshots", 0),
            resumen.get("minutos_vistos_est", 0),
        ]
        ws.append_row(row)
        print(f"[Sheets] ✅ Fila agregada: {resumen.get('nombre')} — {resumen.get('titulo','')[:40]}")
    except Exception as e:
        print(f"[Sheets] Error al escribir: {e}")

# ─── CONFIG ──────────────────────────────────────────────────────────────────

CANALES = [
    {"canal": "lacobraaa",         "nombre": "La Cobra",          "pais": "ARG"},
    {"canal": "agusneta",          "nombre": "Agusneta",          "pais": "ARG"},
    {"canal": "benitosdr",         "nombre": "BenitoSDR",         "pais": "ARG"},
    {"canal": "teodeliaa",         "nombre": "Teo D'Elia",        "pais": "ARG"},
    {"canal": "mikemaquinadelmal", "nombre": "Mike Máq. del Mal", "pais": "MEX"},
    {"canal": "werevertumorro",    "nombre": "Werevertumorro",    "pais": "MEX"},
    {"canal": "goat",              "nombre": "Canal GOAT",        "pais": "BRA"},
    {"canal": "pipesierra",        "nombre": "Pipe Sierra",       "pais": "COL"},
    {"canal": "stagtv",            "nombre": "StagTV",            "pais": "MEX"},
    {"canal": "spiderkongtv",      "nombre": "SpiderKong",        "pais": "MEX"},
]

KEYWORDS_BRA = [
    "brasileir", "brasileirao", "brasileirão", "brazilian serie", "serie a brasil",
    "flamengo", "palmeiras", "corinthians", "santos", "sao paulo", "vasco",
    "botafogo", "cruzeiro", "atletico", "gremio", "inter", "chapecoense",
    "sport", "bragantino", "fortaleza", "bahia", "ceara", "goias",
    "cuiaba", "juventude", "athletico", "america",
]

INTERVALO     = 60
CSV_FILE      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kick_brasileirao_data.csv")
PARTIDOS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kick_partidos_resumen.json")

CSV_COLUMNS = [
    "timestamp", "canal", "nombre", "pais", "estado",
    "viewers_actuales", "peak_sesion", "avg_viewers", "followers",
    "titulo_stream", "categoria", "duracion_min",
]

KICK_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Referer": "https://kick.com/",
}

# ─── ESTADO GLOBAL ───────────────────────────────────────────────────────────

state = {
    "canales": [],
    "brasileirao_activo": False,
    "ultimo_update": None,
    "historial_viewers": {},
    "peaks_sesion": {},
    "sumas_viewers": {},
    "conteo_snapshots": {},
    "sesion_inicio": {},
    "partidos_resumen": [],
    "log": [],
}

_prev_bra: dict = {}  # canal -> titulo (para detectar cuando termina)

# ─── PERSISTENCIA PARTIDOS ───────────────────────────────────────────────────

def load_partidos():
    if os.path.exists(PARTIDOS_FILE):
        with open(PARTIDOS_FILE, encoding="utf-8") as f:
            state["partidos_resumen"] = json.load(f)

def save_partidos():
    with open(PARTIDOS_FILE, "w", encoding="utf-8") as f:
        json.dump(state["partidos_resumen"], f, ensure_ascii=False, indent=2)

def cerrar_sesion(canal, titulo, nombre, pais):
    """Cuando termina un stream de Brasileirao, guarda el resumen con todos los KPIs."""
    peak   = state["peaks_sesion"].get(canal, 0)
    suma   = state["sumas_viewers"].get(canal, 0)
    conteo = state["conteo_snapshots"].get(canal, 1)
    inicio = state["sesion_inicio"].get(canal)
    avg    = round(suma / conteo) if conteo else 0
    ahora  = datetime.now(TZ_AR)
    dur    = int((ahora - inicio).total_seconds() / 60) if inicio else 0

    resumen = {
        "canal":              canal,
        "nombre":             nombre,
        "pais":               pais,
        "titulo":             titulo,
        "fecha":              ahora.strftime("%d/%m/%Y"),
        "hora_inicio":        inicio.strftime("%H:%M") if inicio else "—",
        "hora_fin":           ahora.strftime("%H:%M"),
        "duracion_min":       dur,
        "peak_viewers":       peak,
        "avg_viewers":        avg,
        "total_snapshots":    conteo,
        "minutos_vistos_est": avg * dur,
    }
    state["partidos_resumen"].insert(0, resumen)
    state["partidos_resumen"] = state["partidos_resumen"][:50]
    save_partidos()
    update_excel(resumen)
    update_sheets(resumen)
    log_msg(f"📊 Partido cerrado: {nombre} — peak {peak:,} | avg {avg:,} | {dur}min")

    for d in [state["peaks_sesion"], state["sumas_viewers"],
              state["conteo_snapshots"], state["sesion_inicio"]]:
        d.pop(canal, None)

# ─── CSV ─────────────────────────────────────────────────────────────────────

def init_csv():
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
            csv.DictWriter(f, fieldnames=CSV_COLUMNS).writeheader()
        log_msg(f"CSV creado: {CSV_FILE}")

def save_to_csv(row):
    with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
        csv.DictWriter(f, fieldnames=CSV_COLUMNS).writerow(
            {k: row.get(k, "") for k in CSV_COLUMNS})

# ─── API KICK ────────────────────────────────────────────────────────────────

def fetch_channel(canal):
    try:
        r = _session.get(f"https://kick.com/api/v2/channels/{canal}",
                         headers=KICK_HEADERS, timeout=10)
        return r.json() if r.status_code == 200 else None
    except Exception:
        return None

def parse_channel(data, info):
    canal  = info["canal"]
    nombre = info["nombre"]
    pais   = info["pais"]
    base   = dict(canal=canal, nombre=nombre, pais=pais,
                  estado="offline", viewers_actuales=0,
                  peak_sesion=0, avg_viewers=0, followers=0,
                  titulo_stream="", categoria="", duracion_min=0,
                  es_brasileirao=False,
                  timestamp=datetime.now(TZ_AR).isoformat())

    if not data:
        return base

    base["followers"] = data.get("followersCount") or data.get("followers_count") or 0
    ls = data.get("livestream")
    if not ls:
        return base

    base["estado"]           = "live"
    base["viewers_actuales"] = ls.get("viewer_count") or 0
    base["titulo_stream"]    = ls.get("session_title") or ""
    cats = ls.get("categories") or []
    base["categoria"] = cats[0].get("name") or "" if cats else ""

    texto = (base["titulo_stream"] + " " + base["categoria"]).lower()
    base["es_brasileirao"] = any(kw in texto for kw in KEYWORDS_BRA)

    # Peak
    prev = state["peaks_sesion"].get(canal, 0)
    if base["viewers_actuales"] > prev:
        state["peaks_sesion"][canal] = base["viewers_actuales"]
    base["peak_sesion"] = state["peaks_sesion"].get(canal, base["viewers_actuales"])

    # Acumular avg (solo si es Brasileirao)
    if base["es_brasileirao"]:
        state["sumas_viewers"][canal]    = state["sumas_viewers"].get(canal, 0) + base["viewers_actuales"]
        state["conteo_snapshots"][canal] = state["conteo_snapshots"].get(canal, 0) + 1
        if canal not in state["sesion_inicio"]:
            state["sesion_inicio"][canal] = datetime.now(TZ_AR)
        conteo = state["conteo_snapshots"].get(canal, 1)
        base["avg_viewers"] = round(state["sumas_viewers"][canal] / conteo)

    # Duración
    created = ls.get("created_at") or ls.get("start_time") or ""
    if created:
        try:
            start = datetime.fromisoformat(created.replace("Z", "+00:00"))
            base["duracion_min"] = int((datetime.now(start.tzinfo) - start).total_seconds() / 60)
        except Exception:
            pass

    return base

# ─── LOG ─────────────────────────────────────────────────────────────────────

def log_msg(msg):
    line = f"[{datetime.now(TZ_AR).strftime('%H:%M:%S')}] {msg}"
    print(line)
    state["log"].append(line)
    if len(state["log"]) > 30:
        state["log"].pop(0)

# ─── LOOP ────────────────────────────────────────────────────────────────────

def monitor_loop():
    init_csv()
    load_partidos()
    log_msg("Agente iniciado. Monitoreando canales...")

    while True:
        snapshot = []
        hay_bra  = False

        for info in CANALES:
            canal  = info["canal"]
            raw    = fetch_channel(canal)
            row    = parse_channel(raw, info)
            snapshot.append(row)

            if row["es_brasileirao"]:
                hay_bra = True
                _prev_bra[canal] = row["titulo_stream"]
                save_to_csv(row)
                log_msg(f"🇧🇷 {info['nombre']}: {row['viewers_actuales']:,} viewers (peak {row['peak_sesion']:,})")
            elif canal in _prev_bra:
                # estaba en brasileirao, ahora no → cerrar sesión y guardar KPIs
                cerrar_sesion(canal, _prev_bra.pop(canal), info["nombre"], info["pais"])

            hist = state["historial_viewers"].setdefault(canal, [])
            hist.append(row["viewers_actuales"])
            if len(hist) > 30:
                hist.pop(0)

            time.sleep(0.4)

        state["canales"]            = snapshot
        state["brasileirao_activo"] = hay_bra
        state["ultimo_update"]      = datetime.now(TZ_AR).strftime("%d/%m/%Y %H:%M:%S")

        if not hay_bra:
            log_msg("Sin partido activo. Próxima consulta en 60s.")

        time.sleep(INTERVALO)

# ─── FLASK ───────────────────────────────────────────────────────────────────

app = Flask(__name__)
CORS(app)

@app.route("/api/state")
def api_state():
    csv_rows = 0
    if os.path.exists(CSV_FILE):
        with open(CSV_FILE, encoding="utf-8") as f:
            csv_rows = sum(1 for _ in f) - 1
    return jsonify({
        "canales":            state["canales"],
        "brasileirao_activo": state["brasileirao_activo"],
        "ultimo_update":      state["ultimo_update"],
        "historial":          state["historial_viewers"],
        "partidos_resumen":   state["partidos_resumen"],
        "log":                state["log"],
        "csv_file":           CSV_FILE,
        "csv_rows":           csv_rows,
    })

@app.route("/")
def dashboard():
    return render_template_string(HTML)

# ─── HTML ────────────────────────────────────────────────────────────────────

HTML = r"""
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Brasileirao Monitor</title>
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;700&display=swap" rel="stylesheet">
<style>
:root{--g:#00c244;--g2:#00ff66;--y:#f5c518;--bg:#080c0a;--bg2:#0e1510;--bg3:#141c16;--b:#1e2d22;--text:#d4e8d8;--muted:#4a6650;--ff:'Bebas Neue',sans-serif;--fm:'IBM Plex Mono',monospace;--fb:'IBM Plex Sans',sans-serif}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:var(--fb);min-height:100vh}
body::before{content:'';position:fixed;inset:0;background:repeating-linear-gradient(0deg,transparent,transparent 2px,rgba(0,194,68,.015) 2px,rgba(0,194,68,.015) 4px);pointer-events:none;z-index:1000}
header{display:flex;align-items:center;justify-content:space-between;padding:16px 28px;border-bottom:1px solid var(--b);background:var(--bg2);position:sticky;top:0;z-index:10}
.logo{display:flex;align-items:center;gap:12px}
.logo-icon{width:36px;height:36px;background:var(--g);border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:18px}
.logo h1{font-family:var(--ff);font-size:24px;letter-spacing:2px;color:#fff;line-height:1}
.logo span{font-family:var(--fm);font-size:9px;color:var(--g);letter-spacing:3px;display:block;margin-top:2px}
.pill{font-family:var(--fm);font-size:11px;font-weight:600;letter-spacing:2px;padding:6px 14px;border-radius:100px;text-transform:uppercase;transition:all .4s}
.pill.on{background:rgba(0,194,68,.15);color:var(--g2);border:1px solid var(--g);box-shadow:0 0 12px rgba(0,194,68,.3);animation:pb 2s infinite}
.pill.off{background:rgba(255,60,60,.08);color:var(--muted);border:1px solid var(--b)}
@keyframes pb{0%,100%{box-shadow:0 0 8px rgba(0,194,68,.3)}50%{box-shadow:0 0 20px rgba(0,194,68,.6)}}
.upd{font-family:var(--fm);font-size:10px;color:var(--muted)}
main{padding:24px 28px;max-width:1400px;margin:0 auto}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:22px}
.sc{background:var(--bg2);border:1px solid var(--b);border-radius:10px;padding:16px 20px;position:relative;overflow:hidden;transition:border-color .3s}
.sc::before{content:'';position:absolute;top:0;left:0;width:3px;height:100%;background:var(--g);opacity:.4}
.sc:hover{border-color:var(--g)}
.sl{font-family:var(--fm);font-size:10px;letter-spacing:2px;color:var(--muted);text-transform:uppercase;margin-bottom:6px}
.sv{font-family:var(--ff);font-size:34px;letter-spacing:1px;color:#fff;line-height:1}
.sv.green{color:var(--g2)}
.ss{font-family:var(--fm);font-size:10px;color:var(--muted);margin-top:4px}
.stitle{font-family:var(--ff);font-size:16px;letter-spacing:3px;color:var(--g);text-transform:uppercase;margin-bottom:14px;display:flex;align-items:center;gap:10px}
.stitle::after{content:'';flex:1;height:1px;background:var(--b)}
.cgrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:14px;margin-bottom:24px}
.cc{background:var(--bg2);border:1px solid var(--b);border-radius:12px;padding:18px;transition:all .3s}
.cc.live{border-color:rgba(0,194,68,.3);background:linear-gradient(135deg,var(--bg2),#0a140c)}
.cc.bra{border-color:var(--g);background:linear-gradient(135deg,#091410,#0a1a0d);box-shadow:0 0 18px rgba(0,194,68,.1)}
.cc:hover{transform:translateY(-2px)}
.ctop{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px}
.cnombre{font-family:var(--ff);font-size:20px;letter-spacing:1px;color:#fff}
.chandle{font-family:var(--fm);font-size:10px;color:var(--muted);margin-top:2px}
.badge{font-family:var(--fm);font-size:9px;font-weight:600;letter-spacing:2px;padding:4px 10px;border-radius:100px;text-transform:uppercase}
.b-off{background:rgba(255,255,255,.04);color:var(--muted);border:1px solid var(--b)}
.b-live{background:rgba(0,194,68,.1);color:var(--g2);border:1px solid rgba(0,194,68,.4)}
.b-bra{background:rgba(0,255,102,.15);color:#fff;border:1px solid var(--g2);box-shadow:0 0 8px rgba(0,255,102,.2)}
.cview{font-family:var(--ff);font-size:40px;letter-spacing:1px;color:#fff;line-height:1;margin-bottom:4px}
.cview.z{color:var(--muted);font-size:28px}
.cview.bra{color:var(--g2)}
.cmeta{display:flex;gap:14px;margin-top:6px;flex-wrap:wrap}
.mi{font-family:var(--fm);font-size:10px;color:var(--muted)}
.mi strong{color:var(--text);font-size:11px}
.ctit{margin-top:10px;padding-top:10px;border-top:1px solid var(--b);font-family:var(--fm);font-size:10px;color:var(--muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ctit.bra{color:var(--g)}
.spark{margin-top:8px;height:28px}
canvas{width:100%;height:28px}
.dot{width:7px;height:7px;border-radius:50%;background:var(--g2);display:inline-block;margin-right:5px;box-shadow:0 0 5px var(--g2);animation:bk 1.2s infinite}
@keyframes bk{0%,100%{opacity:1}50%{opacity:.2}}
.csvbar{background:var(--bg3);border:1px solid var(--b);border-radius:8px;padding:10px 18px;display:flex;align-items:center;justify-content:space-between;margin-bottom:20px;font-family:var(--fm);font-size:11px;color:var(--muted)}
.csvbar strong{color:var(--g)}
.bgrid{display:grid;grid-template-columns:1fr 1fr;gap:18px}
.panel{background:var(--bg2);border:1px solid var(--b);border-radius:12px;padding:18px}

/* PARTIDO CARD */
.pc{background:var(--bg3);border:1px solid var(--b);border-radius:10px;padding:16px;margin-bottom:10px}
.pc:last-child{margin-bottom:0}
.pc-top{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px}
.pc-titulo{font-family:var(--fm);font-size:11px;color:var(--g2);font-weight:600}
.pc-meta{font-family:var(--fm);font-size:10px;color:var(--muted)}
.pc-kpis{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:10px}
.kpi{text-align:center;background:var(--bg2);border-radius:6px;padding:10px 6px}
.kv{font-family:var(--ff);font-size:26px;color:#fff;line-height:1}
.kv.peak{color:var(--g2)}
.kv.avg{color:var(--y)}
.kv.dur{color:#7eb8ff}
.kl{font-family:var(--fm);font-size:9px;color:var(--muted);letter-spacing:1px;text-transform:uppercase;margin-top:3px}
.pc-extra{display:flex;gap:16px;flex-wrap:wrap;padding-top:10px;border-top:1px solid var(--b)}
.pe{font-family:var(--fm);font-size:10px;color:var(--muted)}
.pe strong{color:var(--text)}

.log-lines{font-family:var(--fm);font-size:11px;color:var(--muted);line-height:1.8;max-height:320px;overflow-y:auto}
.log-lines .lb{color:var(--g2)}
.log-lines .ls{color:var(--text)}
.empty{font-family:var(--fm);font-size:11px;color:var(--muted);padding:20px 0;text-align:center;line-height:1.6}
@media(max-width:900px){.stats{grid-template-columns:repeat(2,1fr)}.bgrid{grid-template-columns:1fr}main{padding:14px}}
</style>
</head>
<body>
<header>
  <div class="logo">
    <div class="logo-icon">⚽</div>
    <div><h1>BRASILEIRAO MONITOR</h1><span>Kick Streaming Intelligence</span></div>
  </div>
  <div style="display:flex;align-items:center;gap:16px">
    <span class="upd" id="upd">—</span>
    <div class="pill off" id="pill">SIN PARTIDO</div>
  </div>
</header>
<main>
  <div class="stats">
    <div class="sc"><div class="sl">Total Viewers Ahora</div><div class="sv green" id="s-total">0</div><div class="ss">sumados todos los canales live</div></div>
    <div class="sc"><div class="sl">Canales en Vivo</div><div class="sv" id="s-live">0</div><div class="ss">de <span id="s-tc">8</span> monitoreados</div></div>
    <div class="sc"><div class="sl">Peak Máximo (sesión)</div><div class="sv" id="s-peak">0</div><div class="ss" id="s-peak-c">—</div></div>
    <div class="sc"><div class="sl">Registros CSV</div><div class="sv" id="s-csv">0</div><div class="ss">solo partidos detectados</div></div>
  </div>
  <div class="csvbar">
    <span>💾 <strong id="csv-f">kick_brasileirao_data.csv</strong></span>
    <span><strong id="csv-r">0</strong> filas · Solo cuando hay partido activo</span>
  </div>
  <div class="stitle"><span class="dot" id="mdot" style="opacity:.3"></span>Canales Monitoreados</div>
  <div class="cgrid" id="cgrid"></div>
  <div class="bgrid">
    <div class="panel">
      <div class="stitle">📊 KPIs por Partido</div>
      <div id="plist"><div class="empty">Cuando termine un partido, los KPIs<br>quedan acá guardados permanentemente</div></div>
    </div>
    <div class="panel">
      <div class="stitle">📡 Log en Vivo</div>
      <div class="log-lines" id="log"></div>
    </div>
  </div>
  <div class="panel" style="margin:0 1.5rem 2rem;">
    <div class="stitle">📋 Historial Completo de Partidos</div>
    <div style="overflow-x:auto">
    <table id="htable" style="width:100%;border-collapse:collapse;font-family:var(--fm);font-size:12px;color:var(--g1)">
      <thead><tr style="background:var(--g4);color:var(--gr)">
        <th style="padding:8px 12px;text-align:left">Fecha</th>
        <th style="padding:8px 12px;text-align:left">Streamer</th>
        <th style="padding:8px 12px;text-align:left">Partido</th>
        <th style="padding:8px 12px;text-align:right">Peak</th>
        <th style="padding:8px 12px;text-align:right">Avg</th>
        <th style="padding:8px 12px;text-align:right">Duración</th>
        <th style="padding:8px 12px;text-align:right">Min. Vistos</th>
      </tr></thead>
      <tbody id="htbody"></tbody>
    </table>
    </div>
  </div>
</main>
<script>
const fmt = n => Number(n).toLocaleString('es-AR');
function sparkline(id,data){
  const c=document.getElementById(id); if(!c)return;
  const W=c.offsetWidth||280,H=28; c.width=W;c.height=H;
  const ctx=c.getContext('2d'); ctx.clearRect(0,0,W,H);
  if(!data||data.length<2)return;
  const max=Math.max(...data)||1,step=W/(data.length-1);
  ctx.beginPath();
  data.forEach((v,i)=>{const x=i*step,y=H-(v/max)*(H-4)-2;i===0?ctx.moveTo(x,y):ctx.lineTo(x,y);});
  ctx.strokeStyle='rgba(0,255,102,.6)';ctx.lineWidth=1.5;ctx.stroke();
  ctx.lineTo((data.length-1)*step,H);ctx.lineTo(0,H);ctx.closePath();
  ctx.fillStyle='rgba(0,194,68,.07)';ctx.fill();
}
async function refresh(){
  let d; try{const r=await fetch('/api/state');d=await r.json();}catch(e){return;}
  document.getElementById('upd').textContent='Actualizado: '+(d.ultimo_update||'—');
  const pill=document.getElementById('pill'),dot=document.getElementById('mdot');
  if(d.brasileirao_activo){pill.className='pill on';pill.textContent='🔴 PARTIDO EN VIVO';dot.style.opacity='1';}
  else{pill.className='pill off';pill.textContent='SIN PARTIDO';dot.style.opacity='.3';}
  const live=d.canales.filter(c=>c.estado==='live');
  const totalV=live.reduce((s,c)=>s+(c.viewers_actuales||0),0);
  document.getElementById('s-total').textContent=fmt(totalV);
  document.getElementById('s-live').textContent=live.length;
  document.getElementById('s-tc').textContent=d.canales.length;
  document.getElementById('s-csv').textContent=fmt(d.csv_rows);
  document.getElementById('csv-r').textContent=fmt(d.csv_rows);
  document.getElementById('csv-f').textContent=(d.csv_file||'').split('/').pop();
  let maxP=0,maxC='—';
  d.canales.forEach(c=>{if((c.peak_sesion||0)>maxP){maxP=c.peak_sesion;maxC=c.nombre;}});
  document.getElementById('s-peak').textContent=fmt(maxP);
  document.getElementById('s-peak-c').textContent=maxC;

  // Canales
  const grid=document.getElementById('cgrid'); grid.innerHTML='';
  d.canales.forEach(c=>{
    const isBra=c.es_brasileirao,isLive=c.estado==='live',hist=d.historial[c.canal]||[];
    let cc='cc';if(isBra)cc+=' bra';else if(isLive)cc+=' live';
    const badge=isBra?`<span class="badge b-bra">🇧🇷 Brasileirao</span>`:
                 isLive?`<span class="badge b-live"><span class="dot"></span>LIVE</span>`:
                 `<span class="badge b-off">offline</span>`;
    let vc='cview';if(!isLive)vc+=' z';else if(isBra)vc+=' bra';
    const el=document.createElement('div'); el.className=cc;
    el.innerHTML=`
      <div class="ctop">
        <div><div class="cnombre">${c.nombre} <small style="font-size:12px;color:var(--muted)">${c.pais}</small></div>
        <div class="chandle">kick.com/${c.canal}</div></div>${badge}
      </div>
      <div class="${vc}">${isLive?fmt(c.viewers_actuales):'—'}</div>
      <div class="cmeta">
        <div class="mi">Peak <strong>${fmt(c.peak_sesion||0)}</strong></div>
        <div class="mi">Avg <strong>${fmt(c.avg_viewers||0)}</strong></div>
        <div class="mi">Duración <strong>${c.duracion_min||0}min</strong></div>
        <div class="mi">Followers <strong>${fmt(c.followers||0)}</strong></div>
      </div>
      <div class="${isBra?'ctit bra':'ctit'}" title="${c.titulo_stream||''}">${c.titulo_stream||(isLive?c.categoria||'Sin título':'Sin transmisión')}</div>
      <div class="spark"><canvas id="sp-${c.canal}"></canvas></div>`;
    grid.appendChild(el);
    setTimeout(()=>sparkline('sp-'+c.canal,hist),50);
  });

  // KPIs por partido
  const pl=document.getElementById('plist');
  if(!d.partidos_resumen||!d.partidos_resumen.length){
    pl.innerHTML='<div class="empty">Cuando termine un partido, los KPIs<br>quedan acá guardados permanentemente</div>';
  } else {
    pl.innerHTML=d.partidos_resumen.map(p=>`
      <div class="pc">
        <div class="pc-top">
          <div class="pc-titulo">⚽ ${p.titulo||'Partido'}</div>
          <div class="pc-meta">${p.fecha} · ${p.nombre} (${p.pais})</div>
        </div>
        <div class="pc-kpis">
          <div class="kpi"><div class="kv peak">${fmt(p.peak_viewers)}</div><div class="kl">Peak Viewers</div></div>
          <div class="kpi"><div class="kv avg">${fmt(p.avg_viewers)}</div><div class="kl">Avg Viewers</div></div>
          <div class="kpi"><div class="kv dur">${p.duracion_min}</div><div class="kl">Minutos</div></div>
        </div>
        <div class="pc-extra">
          <div class="pe">Inicio <strong>${p.hora_inicio}</strong></div>
          <div class="pe">Fin <strong>${p.hora_fin}</strong></div>
          <div class="pe">Snapshots <strong>${p.total_snapshots}</strong></div>
          <div class="pe">Min. vistos est. <strong>${fmt(p.minutos_vistos_est)}</strong></div>
        </div>
      </div>`).join('');
  }

  // Historial tabla
  const tbody=document.getElementById('htbody');
  if(d.partidos_resumen&&d.partidos_resumen.length){
    tbody.innerHTML=d.partidos_resumen.map((p,i)=>`<tr style="background:${i%2===0?'transparent':'rgba(255,255,255,.03)'}">
      <td style="padding:7px 12px">${p.fecha}</td>
      <td style="padding:7px 12px">${p.nombre} (${p.pais})</td>
      <td style="padding:7px 12px;max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${p.titulo||''}">${p.titulo||'—'}</td>
      <td style="padding:7px 12px;text-align:right;color:#00ff66">${fmt(p.peak_viewers)}</td>
      <td style="padding:7px 12px;text-align:right">${fmt(p.avg_viewers)}</td>
      <td style="padding:7px 12px;text-align:right">${p.duracion_min}min</td>
      <td style="padding:7px 12px;text-align:right">${fmt(p.minutos_vistos_est)}</td>
    </tr>`).join('');
  }

  // Log
  document.getElementById('log').innerHTML=[...d.log].reverse().map(l=>`<div class="${l.includes('🇧🇷')?'lb':l.includes('📺')?'ls':''}">${l}</div>`).join('');
}
refresh(); setInterval(refresh,15000);
</script>
</body>
</html>
"""

# ─── MAIN ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("""
╔══════════════════════════════════════════════════════╗
║   🏆  BRASILEIRAO EN KICK — MONITOR v2              ║
║   Dashboard en vivo + KPIs persistentes              ║
╠══════════════════════════════════════════════════════╣
║   Abrí: http://localhost:5001                        ║
╚══════════════════════════════════════════════════════╝
""")
    t = threading.Thread(target=monitor_loop, daemon=True)
    t.start()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5001)), debug=False, use_reloader=False)


# ─── EXCEL EXPORT (se agrega al final del archivo) ───────────────────────────
# Esta función se llama automáticamente desde cerrar_sesion()

def update_excel(resumen: dict):
    """Crea o actualiza el Excel de KPIs con una fila por partido."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kick_brasileirao_kpis.xlsx")

    COLS = [
        ("Fecha",             "fecha"),
        ("Canal",             "canal"),
        ("Streamer",          "nombre"),
        ("País",              "pais"),
        ("Partido",           "titulo"),
        ("Hora Inicio",       "hora_inicio"),
        ("Hora Fin",          "hora_fin"),
        ("Duración (min)",    "duracion_min"),
        ("Peak Viewers",      "peak_viewers"),
        ("Avg Viewers",       "avg_viewers"),
        ("Snapshots",         "total_snapshots"),
        ("Min. Vistos Est.",  "minutos_vistos_est"),
    ]

    # Colores
    HEADER_BG  = "1A7A3A"
    HEADER_FG  = "FFFFFF"
    ALT_ROW    = "E8F5E9"
    BORDER_CLR = "CCCCCC"

    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cargar o crear workbook
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KPIs por Partido"

        # Headers
        for col_idx, (header, _) in enumerate(COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
            cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        ws.row_dimensions[1].height = 22
        next_row = 2

    # Fila de datos
    for col_idx, (_, key) in enumerate(COLS, 1):
        val  = resumen.get(key, "")
        cell = ws.cell(row=next_row, column=col_idx, value=val)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        if next_row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=ALT_ROW)

    ws.row_dimensions[next_row].height = 18

    # Anchos de columna
    col_widths = [12, 16, 18, 8, 45, 12, 12, 16, 14, 14, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(EXCEL_FILE)
    log_msg(f"📊 Excel actualizado: kick_brasileirao_kpis.xlsx ({next_row - 1} partidos)")
